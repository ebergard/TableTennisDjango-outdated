# -*- coding: utf-8 -*-

import sys
import os
import random
import re
import smtplib
import logging
from email.mime.text import MIMEText
from email.utils import COMMASPACE
from openpyxl import *
from openpyxl.styles import *
from datetime import *
from time import *
from tournament.scripts.constants import *
from tournament.scripts.classes import Participant, Couple, Game, Day


# Create LOG_FOLDER if it doesn't exist
if not os.path.exists(LOG_FOLDER):
    os.makedirs(LOG_FOLDER)

# Set logging
logger = logging.getLogger("log")
logger.setLevel(logging.DEBUG)

fh = logging.FileHandler(os.path.join(LOG_FOLDER, "debug"), mode='w')
formatter = logging.Formatter('%(message)s')
fh.setFormatter(formatter)
logger.addHandler(fh)


def get_number_of_rounds():

    if SINGLE:
        if NUMBER_OF_PARTICIPANTS % 2 == 0:
            return 10
        else:
            return 5
    else:
        if NUMBER_OF_PARTICIPANTS % 4 == 0:
            return 8
        elif NUMBER_OF_PARTICIPANTS % 4 == 2:
            return 4
        else:
            return 2


def get_coefficient():

    return 2 if SINGLE else 4


def generate_couples(games, participants):

    random.shuffle(participants)
    number_of_participants = len(participants)
    couples = []

    if DEBUG:
        logger.debug("Generating couples from: " + str([str(p) for p in participants]))

    for i in range(number_of_participants // 2):

        couple = Couple(participants[i], participants[number_of_participants - i - 1])

        if is_duplicate(games, couple):
            if DEBUG:
                logger.debug("Duplicate couple: " + str(couple))
        else:
            couples.append(couple)

    return couples


def is_duplicate(games, couple):

    if SINGLE:
        for game in games:
            if couple.participant1 == game.player1 and couple.participant2 == game.player2:
                return True
        return False
    else:
        for game in games:
            if couple == game.player1 or couple == game.player2:
                return True
        return False


def games_per_person(participants):

    games_per_person_list = [p.number_of_games for p in participants]
    return games_per_person_list


def games_number_is_equal(participants):

    if len(set(games_per_person(participants))) == 1:
        return True
    else:
        return False


def no_duplicates(couple1, couple2):

    if couple1.participant1 in couple2.participant1.rivals:
        if couple1.participant1.duplicates_count() >= MAX_DUPLICATES2 or \
                couple2.participant1.duplicates_count() >= MAX_DUPLICATES2:
            return False

        if couple1.participant1.duplicate_count(couple2.participant1) >= MAX_DUPLICATES:
            return False
    if couple1.participant2 in couple2.participant1.rivals:
        if couple1.participant2.duplicates_count() >= MAX_DUPLICATES2 or \
                couple2.participant1.duplicates_count() >= MAX_DUPLICATES2:
            return False

        if couple1.participant2.duplicate_count(couple2.participant1) >= MAX_DUPLICATES:
            return False
    if couple1.participant1 in couple2.participant2.rivals:
        if couple1.participant1.duplicates_count() >= MAX_DUPLICATES2 or \
                couple2.participant2.duplicates_count() >= MAX_DUPLICATES2:
            return False

        if couple1.participant1.duplicate_count(couple2.participant2) >= MAX_DUPLICATES:
            return False
    if couple1.participant2 in couple2.participant2.rivals:
        if couple1.participant2.duplicates_count() >= MAX_DUPLICATES2 or \
                couple2.participant2.duplicates_count() >= MAX_DUPLICATES2:
            return False

        if couple1.participant2.duplicate_count(couple2.participant2) >= MAX_DUPLICATES:
            return False

    """
    if couple1.participant1 in couple2.participant1.rivals or \
            couple1.participant1 in couple2.participant2.rivals:
        if couple1.participant1.duplicate_count(couple2.participant1) >= MAX_DUPLICATES or \
                couple1.participant1.duplicate_count(couple2.participant2) >= MAX_DUPLICATES:
            return False

    if couple1.participant2 in couple2.participant1.rivals or \
            couple1.participant2 in couple2.participant2.rivals:
        if couple1.participant2.duplicate_count(couple2.participant1) >= MAX_DUPLICATES or \
                couple1.participant2.duplicate_count(couple2.participant2) >= MAX_DUPLICATES:
            return False
    """
    return True


def generate_games():
    
    # Define variables depending on constants
    number_of_rounds = get_number_of_rounds()
    divider = get_coefficient()
    participants = [Participant(i) for i in range(1, NUMBER_OF_PARTICIPANTS + 1)]
    ###

    games = []
    rounds_counter = 0
    attempts_number = 50
    attempt = 0
    while rounds_counter != number_of_rounds:

        # generate possible couples
        if games_number_is_equal(participants):
            couples_round = generate_couples(games, participants)
        else:
            min_games = min(games_per_person(participants))
            subset_participants = [p for p in participants if p.number_of_games == min_games]

            if len(subset_participants) < divider:
                for p in participants:
                    if p not in subset_participants:
                        subset_participants.append(p)
                        if len(subset_participants) == divider:
                            break

            couples_round = generate_couples(games, subset_participants)

        # generate possible games
        if SINGLE:
            for couple in couples_round:
                games.append(Game(couple.participant1, couple.participant2))
        else:
            if len(couples_round) > 1:
                if games_number_is_equal(participants):
                    games_number_before = len(games)
                    games_number_after = games_number_before
                    while games_number_before == games_number_after:
                        fit_couples_in_game(games, couples_round)
                        games_number_after = len(games)
                else:
                    fit_couples_in_game(games, couples_round)

        # shuffle the list of participants
        random.shuffle(participants)

        if DEBUG:
            for p in participants:
                logger.debug(p.name + ": " + str(p.number_of_games))

        # exit when everybody has the same number of games
        if games_number_is_equal(participants):
            rounds_counter += 1

        # start again if impossible to generate unique couples
        attempt += 1
        if attempt > attempts_number:
            if DEBUG:
                logger.debug("Attempts number for games generation exceeded. Probably bad random. Start again.")

            participants = [Participant(i) for i in range(1, NUMBER_OF_PARTICIPANTS + 1)]
            games = []
            rounds_counter = 0
            attempt = 0

    if DEBUG:
        games_str = ""
        for game in games:
            games_str += str(game.player1) + " vs. " + str(game.player2) + "\n"

        with open(os.path.join(LOG_FOLDER, 'games'), 'w') as fp:
            fp.write(games_str)
    ###########################################################

    # check the same couple in all games
    # games.append(games[-2])
    for i in range(len(games) - 1):
        for j in range(i + 1, len(games)):
            if SINGLE:
                if games[i] == games[j]:
                    sys.exit("Error: the couple is present in more than 1 game: " +
                             str(games[i]) + ", " + str(games[j]))
            else:
                if games[i].player1 == games[j].player1 or games[i].player1 == games[j].player2 or \
                        games[i].player2 == games[j].player1 or games[i].player2 == games[j].player2:
                    sys.exit("Error: the couple is present in more than 1 game: " +
                             str(games[i]) + ", " + str(games[j]))

    # check the number of games per person
    games_per_person_list = [0] * NUMBER_OF_PARTICIPANTS
    for game in games:
        if SINGLE:
            games_per_person_list[game.player1.id - 1] += 1
            games_per_person_list[game.player2.id - 1] += 1
        else:
            games_per_person_list[game.player1.participant1.id - 1] += 1
            games_per_person_list[game.player1.participant2.id - 1] += 1
            games_per_person_list[game.player2.participant1.id - 1] += 1
            games_per_person_list[game.player2.participant2.id - 1] += 1

    # games_per_person_list[0] += 1
    if len(set(games_per_person_list)) != 1:
        sys.exit("Error: players have different number of games: " + str(games_per_person_list))

    return games, games_per_person(participants)[0], participants


def fit_couples_in_game(games, couples_round):

    random.shuffle(couples_round)
    for i in range(len(couples_round) // 2):
        if no_duplicates(couples_round[i], couples_round[len(couples_round) - i - 1]):
            games.append(Game(couples_round[i], couples_round[len(couples_round) - i - 1]))
        else:
            if DEBUG:
                logger.debug("Couples don't fit due to rivals number: " + str(couples_round[i]) + " and " +
                             str(couples_round[len(couples_round) - i - 1]))


def get_time(start_time):
    
    return str(datetime.strptime(start_time, '%H:%M') + timedelta(minutes=TIME_INTERVAL))[11:-3]


def get_dates():
    
    days = [Day(START_DAY)]

    i = 1
    day = datetime.strptime(START_DAY, '%d.%m.%Y')
    while i < NUMBER_OF_DAYS:

        day = day + timedelta(days=1)
        day_str = day.strftime("%a")
        if "Sun" in day_str or "Sat" in day_str:
            if day.strftime("%d.%m.%Y") != "09.06.2018":
                continue
        days.append(Day(day.strftime("%d.%m.%Y")))
        i += 1
    return days


def nobody_plays_this_day(game, day):

    players = []
    if SINGLE:
        players.append(game.player1)
        players.append(game.player2)
    else:
        pass

    for player in players:
        if player in day.participants:
            return False
    return True


def generate_schedule(games):

    # Define variables depending on the constants
    max_games_per_day = len(games) // NUMBER_OF_DAYS
    if len(games) % NUMBER_OF_DAYS != 0:
        max_games_per_day += 1

    days = get_dates()
    coef = get_coefficient()
    ###

    not_set_games = []
    for game in games:
        for i in range(NUMBER_OF_DAYS):
            if len(days[i].participants) < max_games_per_day * coef:
                if nobody_plays_this_day(game, days[i]):
                    if len(days[i].games) == 0:
                        game_time = START_TIME
                    else:
                        game_time = get_time(days[i].games[-1].time)

                    game.set_datetime(days[i].date, game_time)
                    days[i].add_game(game)
                    break
            if i == NUMBER_OF_DAYS - 1:
                not_set_games.append(game)

    if len(not_set_games) == 0:
        schedule = ""
        for day in days:
            for game in day.games:
                schedule += str(game) + "\n"
            schedule += "\n"
        print(schedule)
        with open(os.path.join(LOG_FOLDER, 'schedule'), 'w') as fp:
            fp.write(schedule)

        write_schedule_to_xl(days)

        return True, max_games_per_day, days
    else:
        if DEBUG:
            logger.debug("Not set games:")
            for game in not_set_games:
                logger.debug(str(game))
        return False, None, None


def write_schedule_to_xl(days, sheetname=None, filename="schedule.xlsx"):

    if sheetname is None:
        book = Workbook()
        sheet = book.active
        sheet.title = u"Предварительное расписание"
    else:
        book = load_workbook(os.path.join(LOG_FOLDER, filename.decode("utf-8")))
        sheet = book.create_sheet(sheetname.decode("utf-8"), 0)

    row = 1
    for day in days:
        fill_flag = True
        for i in range(len(day.games)):

            # set data in excel
            sheet.cell(row, 1, day.date)
            sheet.cell(row, 2, day.games[i].time)
            sheet.cell(row, 3, str(day.games[i].player1))
            sheet.cell(row + 1, 3, str(day.games[i].player2))

            # set borders
            thin = Side(border_style="thin", color="000000")
            hair = Side(border_style="hair", color="000000")

            top = Border(top=thin)
            left = Border(left=thin)
            right = Border(right=thin)
            bottom = Border(bottom=thin)
            left2 = Border(left=hair)

            end_col = chr(ord('C') + NUMBER_OF_SETS)
            rows = sheet['A' + str(row):end_col + str(row + 1)]

            for cell in rows[0]:
                cell.border = cell.border + top

            for j in range(0, 4):
                rows[0][j].border = rows[0][j].border + left
                rows[1][j].border = rows[1][j].border + left
            for j in range(4, len(rows[0])):
                rows[0][j].border = rows[0][j].border + left2
                rows[1][j].border = rows[1][j].border + left2
            rows[0][-1].border = rows[0][-1].border + right
            rows[1][-1].border = rows[1][-1].border + right

            for cell in rows[-1]:
                cell.border = cell.border + bottom

            # set background
            if fill_flag:
                fill = PatternFill("solid", fgColor="DEDEDE")
                for _row in rows:
                    for cell in _row:
                        cell.fill = fill
                fill_flag = False
            else:
                fill_flag = True

            row += 2
        row += 1
    book.save(os.path.join(LOG_FOLDER, filename))


def write_games_to_xl(participants, sheetname=u"Игры", filename="schedule.xlsx"):

    book = load_workbook(os.path.join(LOG_FOLDER, filename))
    sheet = book.create_sheet(sheetname, 0)

    participants.sort(key=lambda elem: elem.id)
    row = 1
    for p in participants:
        for game in p.games:
            if SINGLE:
                p1, vs, p2 = str(game).split()[2:5]
                if p.id == int(p2):
                    p1, p2 = p2, p1
                sheet.cell(row, 1, p1)
                sheet.cell(row, 2, vs)
                sheet.cell(row, 3, p2)
            else:
                players = re.findall('\d+', str(game))[5:]
                if p.id == int(players[1]):
                    players[0], players[1] = players[1], players[0]
                if p.id == int(players[2]):
                    players[0], players[2] = players[2], players[0]
                    players[1], players[3] = players[3], players[1]
                if p.id == int(players[3]):
                    players[0], players[3] = players[3], players[0]
                    players[1], players[2] = players[2], players[1]
                sheet.cell(row, 1, players[0])
                sheet.cell(row, 2, players[1])
                sheet.cell(row, 3, "vs.")
                sheet.cell(row, 4, players[2])
                sheet.cell(row, 5, players[3])
            row += 1
        row += 1
    book.save(os.path.join(LOG_FOLDER, filename))


def identify_participants(participants, sheetname, filename="Участники.xlsx"):
    """
    Format of the participants sheet:
    Start row = 2
    Column B: name of a participant
    Column C: email
    Column D: assigned ID
    """
    book = load_workbook(filename.decode("utf-8"))
    sheet = book[sheetname.decode("utf-8")]

    rows = sheet['B2':'D' + str(len(participants) + 1)]
    for row in rows:
        for p in participants:
            if row[2].value == p.id:
                p.set_name(row[0].value.encode("utf-8"))
                p.set_email(row[1].value)
                break


def gather_emails(number_of_participants, sheetname, filename="Участники.xlsx"):
    """
    Format of the participants sheet:
    Start row = 2
    Column B: name of a participant
    Column C: email
    Column D: assigned ID
    """
    book = load_workbook(filename.decode("utf-8"))
    sheet = book[sheetname.decode("utf-8")]
    emails = []

    rows = sheet['C2':'C' + str(number_of_participants + 1)]
    for row in rows:
        emails.append(row[0].value.split("@")[0] + "@appliedtech.ru")

    return emails


def load_results(days, filename="schedule.xlsx"):

    book = load_workbook(os.path.join(LOG_FOLDER, filename))
    sheet = book[u"Расписание"]

    row = 1
    for day in days:
        for i in range(len(day.games)):
            for j in range(4, 4 + NUMBER_OF_SETS):
                value1 = sheet.cell(row, j).value
                value2 = sheet.cell(row + 1, j).value
                if value1 is not None and value2 is not None:
                    if day.games[i].add_result([int(value1), int(value2)]):
                        continue
                    break
                else:
                    break
            row += 2
        row += 1


def write_rating_to_xl(participants, filename="rating.xlsx"):

    if SINGLE:
        highlight_number = 16
    else:
        highlight_number = 3

    # Borders properties
    thin = Side(border_style="thin", color="000000")

    right = Border(right=thin)
    bottom = Border(bottom=thin)
    ###

    book = Workbook()
    sheet = book.active
    sheet.title = u"Текущий рейтинг"

    # Set header
    row = 1
    sheet.cell(row, 1, u"Участник")
    sheet.cell(row, 2, u"Количество выигранных партий")
    sheet.cell(row, 3, u"Разница мячей")

    # Set headers font
    ft = Font(size=12)

    # Set header's borders
    rows = sheet['A' + str(row):'C' + str(row)]
    for cell in rows[0]:
        cell.border = cell.border + right
        cell.border = cell.border + bottom
        cell.font = ft

    participants.sort(key=lambda elem: (elem.win_sets, elem.win_balls), reverse=True)
    for p in participants:

        row += 1

        # set data
        sheet.cell(row, 1, p.name)
        sheet.cell(row, 2, p.win_sets)
        sheet.cell(row, 3, p.win_balls)

        # set borders
        rows = sheet['A' + str(row):'C' + str(row)]

        for cell in rows[0]:
            cell.border = cell.border + right

        # set background
        if row <= highlight_number + 1:
            fill = PatternFill("solid", fgColor="44FABE")
            for _row in rows:
                for cell in _row:
                    cell.fill = fill

    # set bottom border
    rows = sheet['A' + str(row):'C' + str(row)]

    for cell in rows[0]:
        cell.border = cell.border + bottom

    book.save(os.path.join(LOG_FOLDER, filename))


def send_emails_with_schedule(participants):

    me = 'kebergard@appliedtech.ru'

    for p in participants:

        text = "Привет!\nТвои предстоящие игры:\n\n"
        games_list = []
        for game in p.games:
            games_list.append(str(game))
        games_list.sort(key=lambda elem: datetime.strptime(elem[:10], '%d.%m.%Y'))
        text += '\n'.join(games_list)
        text += "\n\nУдачи! :)\n\n" \
                "P.S. Расписание весьма условное, просьба идти на встречу," \
                " если вас просят отыграть какие-то матчи раньше, даже если придётся играть несколько раз за день." \
                " Все игры нужно сыграть до 16:00, 9 июня 2018 г.\n\n"

        if APPTECH:
            msg = MIMEText(text, _charset="utf-8")
            msg['Subject'] = 'Расписание игр'
            msg['From'] = me
            msg['To'] = p.email

            s = smtplib.SMTP('mail.appliedtech.ru')
            s.sendmail(me, [p.email], msg.as_string())
            s.quit()
            sleep(random.randint(1, 4))

        with open(os.path.join(LOG_FOLDER, "sent emails"), 'a') as fp:
            fp.write(p.email + "\n")
            fp.write(text)


def send_email_to_all(recipients, subject, text):

    if APPTECH:
        me = 'kebergard@appliedtech.ru'
        msg = MIMEText(text, _charset="utf-8")
        msg['Subject'] = subject
        msg['From'] = me
        msg['To'] = COMMASPACE.join(recipients)

        s = smtplib.SMTP('mail.appliedtech.ru')
        s.sendmail(me, recipients, msg.as_string())
        s.quit()

    with open(os.path.join(LOG_FOLDER, "sent emails"), 'a') as fp:
        fp.write(str(recipients) + "\n")
        fp.write(text)
